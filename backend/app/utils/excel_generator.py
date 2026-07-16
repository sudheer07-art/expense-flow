from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelGenerator:

    @staticmethod
    def generate_monthly_report(report: dict):

        workbook = Workbook()
        sheet = workbook.active

        sheet.title = "Monthly Report"

        # Title
        sheet["A1"] = "ExpenseFlow Monthly Report"
        sheet["A1"].font = Font(
            bold=True,
            size=16
        )

        # Report Info
        sheet["A3"] = "Month"
        sheet["B3"] = report["month"]

        sheet["A4"] = "Year"
        sheet["B4"] = report["year"]

        sheet["A5"] = "Total Expense"
        sheet["B5"] = report["total_expense"]

        sheet["A6"] = "Transactions"
        sheet["B6"] = report["total_transactions"]

        # Category Header
        sheet["A8"] = "Category"
        sheet["B8"] = "Amount"

        sheet["A8"].font = Font(bold=True)
        sheet["B8"].font = Font(bold=True)

        row = 9

        for item in report["categories"]:

            sheet.cell(
                row=row,
                column=1,
                value=item["category"]
            )

            sheet.cell(
                row=row,
                column=2,
                value=item["total"]
            )

            row += 1

        buffer = BytesIO()

        workbook.save(buffer)

        buffer.seek(0)

        return buffer